import os

import logging

import uvicorn
from aiogram import Bot, Dispatcher, types
from aiogram.fsm.storage.memory import MemoryStorage

from fastapi import FastAPI, Query
from fastapi.responses import FileResponse
from aiogram.types import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    WebAppInfo
)
from pydantic import BaseModel
from starlette.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook


logger = logging.getLogger(__name__)
logging.basicConfig(
    level=logging.INFO,
    format=u'%(filename)s:%(lineno)d #%(levelname)-8s [%(asctime)s] - %(name)s - %(message)s',
)


app = FastAPI()
token = '5957563756:AAFOigY3zrgc8ooqUi2w5WjUe2D2t8lxtX8'
bot = Bot(token=token, parse_mode="HTML")
storage = MemoryStorage()
dp = Dispatcher(storage=storage)
# CODASSISTANT_URL = os.getenv('CODASSISTANT_URL')
# HOST = os.getenv('HOST')
# PREFIX = os.getenv('PREFIX')
WEBHOOK_URL = F'95.31.19.28/webhook'
HTML_URL = f'95.31.19.28/scpass/'
HTML_CAR = f'95.31.19.28/scpass/car'

app.mount("/static", StaticFiles(directory="../src/static"), name="static")

# Загрузка данных из файла Excel
wb = load_workbook('../src/static/cars.xlsx')
sheet = wb.active

# # Проверка данных из файла Excel
# for row in sheet.iter_rows(values_only=True):
#     logging.info(row[0].lower())


@app.get("/", response_class=HTMLResponse)
def root():
    return FileResponse("../src/index.html")


@app.get("/car", response_class=HTMLResponse)
def root():
    return FileResponse("../src/car.html")


@app.get("/car_models")
async def suggest(query: str = Query(...)):
    suggestions = []
    logging.info("%s %s", 1111, query)
    # Поиск подходящих значений
    for row1 in sheet.iter_rows(values_only=True):
        if query.lower() in row1[0].lower():
            suggestions.append(row1[0])
    logging.info("%s %s", 33333, suggestions)
    return JSONResponse(content=suggestions)


class WebInterface(BaseModel):
    date: str
    time_from: str
    time_to: str
    fio_worker: str
    purpose_visit: str | None
    fio_visitor: list
    car_info: list
    id: int
    first_name: str
    form_width: int
    form_height: int


class WebCarInterface(BaseModel):
    car_info: list
    id: int
    first_name: str
    form_width: int
    form_height: int


@app.post("/webhook")
async def bot_webhook(update: WebInterface | WebCarInterface | dict):
    print(update)
    if isinstance(update, WebInterface):
        chat_id = update.id
        date = update.date
        time_from = update.time_from
        time_to = update.time_to
        fio_worker = update.fio_worker
        purpose_visit = update.purpose_visit
        fio_visitor = update.fio_visitor
        car_info = update.car_info
        first_name = update.first_name
        form_width = update.form_width
        form_height = update.form_height
        await bot.send_message(chat_id=chat_id, text=f'Здравствуйте {first_name} ваш id: {chat_id}\n'
                                                     f'Вы ввели: {date, time_from, time_to, fio_worker, purpose_visit}'
                                                     f'{fio_visitor, car_info}\nширина - {form_width}\n'
                                                     f'высота - {form_height}')
    elif isinstance(update, WebCarInterface):
        chat_id = update.id
        car_info = update.car_info
        first_name = update.first_name
        form_width = update.form_width
        form_height = update.form_height
        await bot.send_message(chat_id=chat_id, text=f'Здравствуйте {first_name} ваш id: {chat_id}\n'
                                                     f'{car_info}\nширина - {form_width}\n'
                                                     f'высота - {form_height}')

    elif isinstance(update, dict):
        telegram_update = types.Update(**update)
        user_id = telegram_update.message.from_user.id
        button = InlineKeyboardButton(text='обычный', web_app=WebAppInfo(url=HTML_URL))
        car = InlineKeyboardButton(text='на машину', web_app=WebAppInfo(url=HTML_CAR))
        reply_markup = InlineKeyboardMarkup(resize_keyboard=True, inline_keyboard=[[button], [car]])
        await bot.send_message(chat_id=user_id, reply_markup=reply_markup, text='Привет! Я WebApp бот перейди по '
                                                                                'ссылке ниже:')
    return 'ok'


if __name__ == "__main__":
    uvicorn.run(
        app=f"{__name__}:app",
        host="0.0.0.0",
        port=80,
        workers=10,
        # root_path=PREFIX,
        forwarded_allow_ips="*",
        proxy_headers=True,
        limit_concurrency=10,
    )
